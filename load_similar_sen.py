# -*- coding: utf-8 -*-
"""
Created on Thu Mar 16 09:28:09 2017

@author: qxh
"""

from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
import json
import pickle

#导入由陈双判断的百度相似问句,
#输入格式第一列、第二列都是句子，第九列是标注的相似性
#返回格式list['sen1','sen2','sim']
def load_excel_similar_sen(excel_name, first_row=-1):
      
    #读取excel2007文件  
    wb = load_workbook(excel_name)  
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[0])  
    
    sen_list = []
    if first_row == -1:
        max_rows = ws.max_row()
    else:        
        max_rows = min (ws.max_row(), first_row)
    #把数据存到列表中  
    for rx in range(max_rows):  
        dic_sen={}
        dic_sen['sen1'] = ws.cell(row = rx+1,column = 1).value
        dic_sen['sen2'] = ws.cell(row = rx+1,column = 2).value  
        dic_sen['sim'] = ws.cell(row = rx+1,column = 9).value  
        sen_list.append(dic_sen)
    return sen_list


def load_excel_para_sen(excel_name, first_row=-1):
    '''
    导入评测的相似问句。输入格式 改写句，原句，答案。
    :param excel_name:
    :param first_row:
    :return: [{'ground_truth':str, 'paraphrase':str，'answer':str}]
    '''
    # 读取excel2007文件
    wb = load_workbook(excel_name)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])

    sen_list = []
    if first_row == -1:
        max_rows = ws.max_row
    else:
        max_rows = min(ws.max_row, first_row)
    # 把数据存到列表中
    for rx in range(max_rows):
        dic_sen = {}
        dic_sen['paraphrase'] = ws.cell(row=rx + 1, column=1).value
        dic_sen['ground_truth'] = ws.cell(row=rx + 1, column=2).value
        dic_sen['answer'] = ws.cell(row=rx + 1, column=3).value
        sen_list.append(dic_sen)
    return sen_list

#导入由算法组人员改写的知乎相似问句
#输入格式：第一例是原问题，后面二列后全是相似问题
#生成格式：list[dic{'原问题作为key'}:list[相似问题的list]]
def load_excel_similar_sen_modify(excel_name, first_row=-1):
      
    max_col = 10
    #读取改写后的excel文件  
    wb = load_workbook(excel_name)  
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[0])  
    
    sen_list = []
    if first_row == -1:
        max_rows = ws.max_row()
    else:        
        max_rows = min(ws.max_row(), first_row)
        
    #把数据存到列表中  
    for rx in range(max_rows):  
        dic_sen={}
        
        key = ws.cell(row = rx+1,column = 1).value
        if key.split() == "":
            continue
        else:
            value = []
            for c in range(2,max_col):
                v = ws.cell(row = rx+1,column = c).value
                if v.split() == "":
                    continue
                else:
                    value.append(v)
        dic_sen[key] = value
        sen_list.append(dic_sen)
    return sen_list
    

def load_excel_similar_sen_modify_to_para(excel_name, first_row=-1):
    '''
    #导入相似问句
    #输入格式：第一例是原问题，后面二列后全是相似问题
    #生成格式：list[dic{'ground_truth:str','paraphrase:str','candidates:()'}]
    '''
      
    max_col = 10
    #读取改写后的excel文件  
    wb = load_workbook(excel_name)  
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[0])  
    
    sen_list = []
    if first_row == -1:
        max_rows = ws.max_row()
    else:        
        max_rows = min (ws.max_row(), first_row)

    #把数据存到列表中  
    candi_dic = {}
    for rx in range(1,max_rows):  #excel从1开始的
        key = ws.cell(row = rx,column = 1).value
        if len(key) < 2 or key == 'None':
            continue
        else:
            candi_dic[key] = key 
            for c in range(2,max_col):
                v = str(ws.cell(row = rx+1,column = c).value)
                if len(v) < 2 or v == 'None':
                    continue
                else:
                    sen_list.append({'ground_truth':key, 'paraphrase': str(v)})
                    
    candi = list(candi_dic.keys())
    candi = [[row, 1] for row in candi]
    sen_list = [{'ground_truth': row['ground_truth'], 'paraphrase': row['paraphrase'], 'candidates': candi} for row in sen_list]
                    
    return sen_list
    
def load_excel_similar_sen_modify_to_para_full(excel_name, first_row=-1):
    '''
    #特殊定制。导入相似问句.负样本的gr为‘’
    #输入格式：第一例是原问题，2-4列后全是相似问题,5-10列为非相似问题.
    #生成格式：list[dic{'ground_truth:str','paraphrase:str','candidates:list(str)'}]
    '''
      
    max_col = 12
    #读取改写后的excel文件  
    wb = load_workbook(excel_name)  
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[1])  
    
    sen_list = []
    if first_row == -1:
        max_rows = ws.max_row()
    else:        
        max_rows = min (ws.max_row(), first_row)

    #把数据存到列表中  
    candi_dic = {}
    for rx in range(1,max_rows):     #excel从1开始的,加了一列标签    
        key = ws.cell(row = rx+1,column = 1).value
        if len(key) < 2 or key == 'None':
            continue
        else:
            candi_dic[key] = key 
            for c in range(2,max_col):
                v = str(ws.cell(row = rx+1,column = c).value)
                if len(v) < 2 or v == 'None':
                    continue
                if c > 6:
                    sen_list.append({'ground_truth':'', 'paraphrase': str(v)})
                else:
                    sen_list.append({'ground_truth':key, 'paraphrase': str(v)})
                    
    candi = list(candi_dic.keys())
    candi = [[row, 1] for row in candi]
    sen_list = [{'ground_truth': row['ground_truth'], 'paraphrase': row['paraphrase'], 'candidates': candi} for row in sen_list]
                    
    return sen_list

#导入由算法组人员改写的知乎相似问句
#输入格式：侯聪生成的json格式文件，list[dic('original_question'， ‘paraphrase’)]
#生成格式：list[dic{'原问题作为key'}:list[相似问题的list]]
def load_json_similar_sen(json_name): 
    with open(json_name, 'r', encoding='utf8') as file:
        data = json.load(file)
    
    sen_dic = {}
    for i in range(len(data)):
        sen1 = data[i]['original_question']
        if sen1 in sen_dic.keys():
            sen_dic[sen1].append(data[i]['paraphrase'])
        else:
            sen_dic[sen1] = []
            sen_dic[sen1].append(data[i]['paraphrase'])
    
    return sen_dic


#导入相似问句--excel中，
#excel中第一例是原问题，二列是相似问题
#返回格式list['sen1','sen2']
def load_excel_similar_sen_cookbook_test(excel_name, first_row=-1):
      
    #读取excel2007文件  
    wb = load_workbook(excel_name)  
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[0])  
    
    sen_list = []
    if first_row == -1:
        max_rows = ws.max_row()
    else:        
        max_rows = min (ws.max_row(), first_row)
    #把数据存到列表中  
    for rx in range(max_rows):  
        dic_sen={}
        dic_sen['sen1'] = ws.cell(row = rx+1,column = 1).value
        dic_sen['sen2'] = ws.cell(row = rx+1,column = 2).value   
        sen_list.append(dic_sen)
    return sen_list

#导入相似问句--excel中，
#excel中第一例是原问题，二列是相似问题
#生成格式：list[dic{'原问题作为key'}:list[相似问题的list]]
def load_similar_sen_cookbook_test(excel_name, first_row=-1):
    data = load_excel_similar_sen_cookbook_test(excel_name, first_row)
    sen_dic = {}
    for i in range(len(data)):
        sen_or = data[i]['sen2']
        if sen_or in sen_dic.keys():
            sen_dic[sen_or].append(data[i]['sen1'])
        else:
            sen_dic[sen_or] = []
            sen_dic[sen_or].append(data[i]['sen1'])
    
    return sen_dic

def write_list_dic_to_excel(list_dic_name, excel_name='list_dic_to_excel.xlsx'):

    wb=Workbook()
    ws=wb.active
    
    keys = list(list_dic_name[0].keys())
    ws.append(keys)
    
    for i in range(len(list_dic_name)):
        row_dic = []
        for j in keys:
            row_dic.append(list_dic_name[i][j])
        ws.append(row_dic)
    
    wb.save(excel_name)    
    
if __name__ == '__main__':
    
#    excel_path = './result/cookbook_test.xlsx'
#    res = load_similar_sen_cookbook_test(excel_path)
#    js_path = './knowledge/paraphrase_data_new.json'
#    res = load_json_similar_sen(js_path)
    
    # data = load_excel_similar_sen_modify_to_para_full('./liguang.xlsx')
    # sx = [{'ground_truth':row['ground_truth'], 'paraphrase':row['paraphrase']} for row in data]
    # with open('./result/liguang_new', 'wb') as file:
    #     pickle.dump(data, file)
    
#    with open('./result/liguang_old', 'rb') as file:
#        xxdata = pickle.load(file)

    # xx = load_excel_para_sen(excel_name='D:\svn_algorithm\standard\测试平台算法评测\data\测试数据\海关.xlsx', first_row=-1)
    # yy = 0

    # xx = {"1":1, "2":2}
    # yy = []
    # yy.append(xx)
    # write_list_dic_to_excel(yy)

    path = 'D:/evaluation/data/测试数据/finance8000_test.xlsx'
    data = load_excel_para_sen(path, -1)
    print(data)